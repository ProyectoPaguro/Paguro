from __future__ import annotations

import os
import io
from datetime import datetime, date, time as dtime  

def rango_hoy():
    hoy = date.today()
    inicio = datetime.combine(hoy, dtime.min)  
    fin    = datetime.combine(hoy, dtime.max)   
    return inicio, fin
   
import time as epoch_time                           
from pathlib import Path

from flask import (
    Flask, render_template, request, redirect, url_for,
    send_file, flash, abort
)
from flask_sqlalchemy import SQLAlchemy
from flask_login import (
    UserMixin, LoginManager, login_user, login_required,
    logout_user, current_user
)
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from sqlalchemy import func, text, or_

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from flask import Flask, render_template, request, redirect, url_for, send_file, flash, abort, jsonify


import os
from flask import Flask
from flask_sqlalchemy import SQLAlchemy

app = Flask(__name__, instance_relative_config=True)

from zoneinfo import ZoneInfo
from datetime import timezone, time as dtime, datetime

LOCAL_TZ_NAME = os.environ.get("LOCAL_TZ", "America/Bogota")
LOCAL_TZ = ZoneInfo(LOCAL_TZ_NAME)
UTC = ZoneInfo("UTC")

def to_local(dt):
    """Convierte un datetime guardado en UTC (naive) a hora local."""
    if not dt:
        return None
   
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=UTC)
    return dt.astimezone(LOCAL_TZ)

def fmt_local(dt, fmt="%Y-%m-%d %H:%M"):
    x = to_local(dt)
    return x.strftime(fmt) if x else ""

@app.template_filter("localdt")
def localdt_filter(dt, fmt="%Y-%m-%d %H:%M"):
    return fmt_local(dt, fmt)


BODEGAS_FIJAS = [b.strip() for b in os.getenv("BODEGAS", "Tocancipa,Bogotá Prado,Casa Balsa,Ferias").split(",") if b.strip()]

def listar_bodegas():
    """Devuelve bodegas conocidas: fijas + las que existan en BD."""
    otras = [b[0] for b in db.session.query(Producto.bodega).distinct().all() if b[0]]
   
    vistas = []
    for b in BODEGAS_FIJAS + otras:
        if b and b not in vistas:
            vistas.append(b)
    return vistas

from zoneinfo import ZoneInfo

LOCAL_TZ = ZoneInfo(os.getenv("LOCAL_TZ", "America/Bogota"))  
UTC = ZoneInfo("UTC")

def hoy_local_a_utc_bounds():
    """Devuelve (inicio_utc_naive, fin_utc_naive) del día local actual."""
    ahora_local = datetime.now(LOCAL_TZ)
    ini_local = datetime.combine(ahora_local.date(), dtime.min, tzinfo=LOCAL_TZ)
    fin_local = datetime.combine(ahora_local.date(), dtime.max, tzinfo=LOCAL_TZ)
    ini_utc = ini_local.astimezone(UTC).replace(tzinfo=None)  
    fin_utc = fin_local.astimezone(UTC).replace(tzinfo=None)
    return ini_utc, fin_utc
#
_env_secret = os.environ.get("SECRET_KEY")
if not _env_secret:
    try:
        _env_secret = os.urandom(24).hex()
    except Exception:
        _env_secret = "dev-secret-key-change-me"

app.config["SECRET_KEY"] = _env_secret
app.secret_key = _env_secret


print(">>> SECRET_KEY set? ->", bool(app.config.get("SECRET_KEY")))


DATA_DIR = os.environ.get("DATA_DIR", "/var/data")

if os.path.isdir(DATA_DIR) and os.access(DATA_DIR, os.W_OK):
    os.makedirs(DATA_DIR, exist_ok=True)
    db_path = os.path.join(DATA_DIR, "database.db")
else:
    os.makedirs(app.instance_path, exist_ok=True)
    db_path = os.path.join(app.instance_path, "database.db")

app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///" + db_path
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

print(">>> DB file:", db_path)

db = SQLAlchemy(app)

class CategoriaInsumo(db.Model):
    __tablename__ = 'categoria_insumo'
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(80), nullable=False, unique=True)
    descripcion = db.Column(db.String(200))  # ✅ agregado
    insumos = db.relationship('Insumo', backref='categoria', lazy=True)

    def __repr__(self):
        return f'<CategoriaInsumo {self.nombre}>'




class Insumo(db.Model):
    __tablename__ = 'insumo'
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100), nullable=False)
    unidad = db.Column(db.String(50))
    cantidad_actual = db.Column(db.Float, default=0)
    bodega = db.Column(db.String(100))
    categoria_id = db.Column(db.Integer, db.ForeignKey('categoria_insumo.id'))



class Movimiento(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    tipo = db.Column(db.String(20), nullable=False)  
    cantidad = db.Column(db.Float, nullable=False)
    fecha = db.Column(db.DateTime, default=datetime.utcnow)
    insumo_id = db.Column(db.Integer, db.ForeignKey('insumo.id'))
    insumo = db.relationship('Insumo', backref=db.backref('movimientos', lazy=True))


class Usuario(UserMixin, db.Model):
    __tablename__ = 'usuarios'
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100), nullable=False)
    email  = db.Column(db.String(100), unique=True, nullable=False)
    password = db.Column(db.String(200), nullable=False)
    rol = db.Column(db.String(20), nullable=False, default='operario')
   
class RegistroPulido(db.Model):
    __tablename__ = 'registros_pulido'

    id = db.Column(db.Integer, primary_key=True)
    fecha = db.Column(db.Date, default=date.today, nullable=False)

    # quién pulió
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuarios.id'), nullable=False)
    usuario = db.relationship('Usuario', backref=db.backref('registros_pulido', lazy=True))

    # info del producto pulido
    producto = db.Column(db.String(120), nullable=False)
    acabado = db.Column(db.String(120))
    cantidad = db.Column(db.Integer, default=1, nullable=False)
    categoria_id = db.Column(db.Integer, db.ForeignKey('categoria_produccion.id'))
    estado = db.Column(db.String(20), default='pulido', nullable=False)
    observaciones = db.Column(db.Text)

class CategoriaProduccion(db.Model):
    __tablename__ = 'categoria_produccion'
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(80), nullable=False, unique=True)

    productos_produccion = db.relationship('Produccion', backref='categoria_produccion', lazy=True)
    productos = db.relationship('Producto', backref='categoria_producto', lazy=True)



class Produccion(db.Model):
    __tablename__ = 'produccion'
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100), nullable=False)
    cantidad = db.Column(db.Float, default=0)
    unidad = db.Column(db.String(50), nullable=False)
    categoria_id = db.Column(db.Integer, db.ForeignKey('categoria_produccion.id'))

class Producto(db.Model):
    __tablename__ = 'producto'
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(120), nullable=False)
    acabado = db.Column(db.String(120), nullable=False)
    cantidad_actual = db.Column(db.Float, default=0)
    bodega = db.Column(db.String(120), nullable=False)
    categoria_id = db.Column(db.Integer, db.ForeignKey('categoria_produccion.id'))


class ProdMovimiento(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    tipo = db.Column(db.String(20), nullable=False)    
    cantidad = db.Column(db.Float, nullable=False)
    fecha = db.Column(db.DateTime, default=datetime.utcnow)
    bodega = db.Column(db.String(120))                
    producto_id = db.Column(db.Integer, db.ForeignKey('producto.id'))
    producto = db.relationship('Producto', backref=db.backref('movimientos', lazy=True))

class Transferencia(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    producto_nombre = db.Column(db.String(120))
    acabado = db.Column(db.String(120))
    cantidad = db.Column(db.Float)
    origen = db.Column(db.String(120))
    destino = db.Column(db.String(120))
    fecha = db.Column(db.DateTime, default=datetime.utcnow)



class Tarea(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    tipo = db.Column(db.String(20), nullable=False)      
    producto = db.Column(db.String(200), nullable=False)
    acabado = db.Column(db.String(120), nullable=False)
    imagen = db.Column(db.String(255))                   
    fecha = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)

    
    completada = db.Column(db.Boolean, default=False, nullable=False)
    completada_en = db.Column(db.DateTime)
    completada_por_id = db.Column(db.Integer, db.ForeignKey('usuarios.id'))
    completada_por = db.relationship('Usuario', foreign_keys=[completada_por_id])


login_manager = LoginManager()
login_manager.login_view = 'login'
login_manager.init_app(app)

@login_manager.user_loader
def load_user(user_id):
    return db.session.get(Usuario, int(user_id))


with app.app_context():
    db.create_all()
    try:
        cols = db.session.execute(text("PRAGMA table_info(prod_movimiento);")).fetchall()
        colnames = [c[1] for c in cols]
        if 'bodega' not in colnames:
            db.session.execute(text("ALTER TABLE prod_movimiento ADD COLUMN bodega VARCHAR(120);"))
            db.session.commit()
    except Exception as e:
        print("Aviso 'bodega en prod_movimiento':", e)
    try:
        cols = db.session.execute(text("PRAGMA table_info(usuarios);")).fetchall()
        colnames = [c[1] for c in cols]
        if 'rol' not in colnames:
            db.session.execute(
                text("ALTER TABLE usuarios ADD COLUMN rol VARCHAR(20) NOT NULL DEFAULT 'operario';")
            )
            db.session.commit()
    except Exception as e:
        print("Aviso 'rol':", e)

def require_roles(*roles):
    def deco(fn):
        from functools import wraps
        @wraps(fn)
        def wrapper(*args, **kwargs):
            if not current_user.is_authenticated:
                return login_manager.unauthorized()
            if current_user.rol not in roles:
                abort(403)
            return fn(*args, **kwargs)
        return wrapper
    return deco


def _save_task_image(file_storage):
    """La carga de imagen está deshabilitada. Siempre devolver None."""
    return None


@app.route('/favicon.ico')
def favicon():
    return ('', 204)

@app.route('/')
@login_required
def dashboard():
    total_insumos = Insumo.query.count()
    total_productos = Producto.query.count()

    ini_hoy_utc, fin_hoy_utc = hoy_local_a_utc_bounds()

    movs_hoy = (Movimiento.query
                .filter(Movimiento.fecha >= ini_hoy_utc,
                        Movimiento.fecha <= fin_hoy_utc)
                .count())

    UMBRAL = 5
    bajos = Insumo.query.filter(Insumo.cantidad_actual < UMBRAL).count()

    tareas_fundir_pend = (Tarea.query
        .filter(Tarea.tipo=='fundir', Tarea.completada.is_(False))
        .order_by(Tarea.id.desc()).all())

    tareas_pulir_pend = (Tarea.query
        .filter(Tarea.tipo=='pulir', Tarea.completada.is_(False))
        .order_by(Tarea.id.desc()).all())

    tareas_fundir_comp = (Tarea.query
        .filter(Tarea.tipo=='fundir', Tarea.completada.is_(True))
        .filter(Tarea.completada_en >= ini_hoy_utc,
                Tarea.completada_en <= fin_hoy_utc)
        .order_by(Tarea.completada_en.desc().nullslast()).all())

    tareas_pulir_comp = (Tarea.query
        .filter(Tarea.tipo=='pulir', Tarea.completada.is_(True))
        .filter(Tarea.completada_en >= ini_hoy_utc,
                Tarea.completada_en <= fin_hoy_utc)
        .order_by(Tarea.completada_en.desc().nullslast()).all())

    # ✅ REGISTROS DE PULIDO

    # 1) Lo que el usuario actual ha registrado hoy
    registros_pulido_hoy_usuario = (
        RegistroPulido.query
        .filter(
            RegistroPulido.fecha == date.today(),
            RegistroPulido.usuario_id == current_user.id
        )
        .order_by(RegistroPulido.id.desc())
        .all()
    )

    # 2) Todos los registros pendientes (estado="pulido")
    registros_pulido_pendientes = (
        RegistroPulido.query
        .filter(RegistroPulido.estado == "pulido")
        .order_by(RegistroPulido.fecha.desc(), RegistroPulido.id.desc())
        .all()
    )

    categorias_produccion = CategoriaProduccion.query.order_by(CategoriaProduccion.nombre.asc()).all()

    return render_template(
    'dashboard.html',
    total_insumos=total_insumos,
    total_productos=total_productos,
    movs_hoy=movs_hoy,
    bajos=bajos,
    umbral=UMBRAL,
    tareas_fundir_pend=tareas_fundir_pend,
    tareas_fundir_comp=tareas_fundir_comp,
    tareas_pulir_pend=tareas_pulir_pend,
    tareas_pulir_comp=tareas_pulir_comp,
    registros_pulido_hoy_usuario=registros_pulido_hoy_usuario,
    registros_pulido_pendientes=registros_pulido_pendientes,
    categorias_produccion=categorias_produccion   # <-- NUEVO
)


@app.post("/pulido/<int:registro_id>/terminar")
@login_required
@require_roles("admin")   # solo admin; luego puedes cambiarlo si quieres
def pulido_terminar(registro_id):
    reg = RegistroPulido.query.get_or_404(registro_id)

    if reg.estado == "terminado":
        flash("Este registro ya estaba marcado como terminado.", "info")
        return redirect(url_for("dashboard"))

    # --- Buscar o crear el producto en producción ---
    # Puedes cambiar "Tocancipa" por la bodega que uses para producción final
    BODEGA_PRODUCCION = "Tocancipa"

    prod = (
        Producto.query
        .filter_by(
            nombre=reg.producto,
            acabado=reg.acabado,
            bodega=BODEGA_PRODUCCION
        )
        .first()
    )

    if not prod:
        # Si no existe el producto en esa bodega, lo creamos
        prod = Producto(
    nombre=reg.producto,
    acabado=reg.acabado,
    cantidad_actual=0,
    bodega=BODEGA_PRODUCCION,
    categoria_id=reg.categoria_id   # 👈 USAMOS LA CATEGORÍA DEL EMPLEADO
)

        db.session.add(prod)
        db.session.flush()  # para que tenga id

    # --- Aumentar stock en producción ---
    prod.cantidad_actual += reg.cantidad

    # --- Registrar movimiento de producción (Entrada) ---
    mov = ProdMovimiento(
        tipo="Entrada",
        cantidad=reg.cantidad,
        producto=prod,
        fecha=datetime.utcnow(),
        bodega=prod.bodega
    )
    db.session.add(mov)

    # --- Marcar el registro de pulido como terminado ---
    reg.estado = "terminado"

    db.session.commit()

    flash(
        f"Pulido terminado: +{reg.cantidad} de '{reg.producto}' ({reg.acabado}) en {prod.bodega}.",
        "success"
    )
    return redirect(url_for("dashboard"))


@app.post("/tareas/agregar")
@login_required
@require_roles("admin")
def tareas_agregar():
    tipo = (request.form.get("tipo") or "").strip().lower()  
    producto = (request.form.get("producto") or "").strip()
    acabado = (request.form.get("acabado") or "").strip()
    if not tipo or not producto or not acabado:
        flash("Faltan datos para crear la tarea.", "warning")
        return redirect(url_for("dashboard"))

    imagen_rel = None
    if tipo == "fundir":
        imagen_rel = _save_task_image(request.files.get("imagen"))

    t = Tarea(
        tipo=tipo,
        producto=producto,
        acabado=acabado,
        imagen=imagen_rel,
        fecha=datetime.utcnow(),
        completada=False,
    )
    db.session.add(t)
    db.session.commit()
    flash("Tarea añadida ✅", "success")
    return redirect(url_for("dashboard"))

@app.context_processor
def inject_active_path():
    from flask import request
    p = getattr(request, "path", "")
    return {"active_path": p, "ap": p}

@app.post("/tareas/<int:tarea_id>/completar")
@login_required
def tareas_completar(tarea_id):
    t = Tarea.query.get_or_404(tarea_id)
    if not t.completada:
        t.completada = True
        t.completada_en = datetime.utcnow()
        t.completada_por_id = current_user.id
        db.session.commit()
        flash("Tarea completada ✔️", "success")
    return redirect(url_for("dashboard"))

@app.post("/tareas/<int:tarea_id>/reabrir")
@login_required
@require_roles("admin")
def tareas_reabrir(tarea_id):
    t = Tarea.query.get_or_404(tarea_id)
    if t.completada:
        t.completada = False
        t.completada_en = None
        t.completada_por_id = None
        db.session.commit()
        flash("Tarea reabierta ↩️", "info")
    return redirect(url_for("dashboard"))

@app.post("/tareas/<int:tarea_id>/eliminar")
@login_required
@require_roles("admin")
def tareas_eliminar(tarea_id):
    t = Tarea.query.get_or_404(tarea_id)
    if t.imagen:
        try:
            os.remove(os.path.join(app.static_folder, t.imagen))
        except Exception:
            pass
    db.session.delete(t)
    db.session.commit()
    flash("Tarea eliminada 🗑️", "warning")
    return redirect(url_for("dashboard"))

@app.get('/tareas/historial')
@login_required
def historial_tareas():
    tipo = (request.args.get('tipo') or '').lower()            
    estado = (request.args.get('estado') or '').lower()       
    desde = request.args.get('desde')                          
    hasta = request.args.get('hasta')

    q = Tarea.query

    if tipo in ('fundir', 'pulir'):
        q = q.filter(Tarea.tipo == tipo)

    if estado in ('pendiente', 'completada'):
        q = q.filter(Tarea.completada.is_(estado == 'completada'))

    if desde:
        try:
            d = datetime.strptime(desde, "%Y-%m-%d")
            q = q.filter(Tarea.fecha >= d)
        except ValueError:
            flash('Fecha "desde" inválida', 'danger')

    if hasta:
        try:
            h = datetime.strptime(hasta, "%Y-%m-%d")
            h = h.replace(hour=23, minute=59, second=59)
            q = q.filter(Tarea.fecha <= h)
        except ValueError:
            flash('Fecha "hasta" inválida', 'danger')

    tareas = q.order_by(Tarea.fecha.desc(), Tarea.id.desc()).all()
    return render_template('historial_tareas.html',
                           tareas=tareas,
                           sel_tipo=tipo, sel_estado=estado,
                           sel_desde=desde or '', sel_hasta=hasta or '')

@app.post("/pulido/registrar")
@login_required
def registrar_pulido():
    # Si luego quieres limitar solo a cierto rol:
    # if current_user.rol != 'operario':
    #     abort(403)

    producto = (request.form.get('producto') or '').strip()
    acabado = (request.form.get('acabado') or '').strip()
    cantidad_str = (request.form.get('cantidad') or '1').strip()
    observaciones = (request.form.get('observaciones') or '').strip()
    categoria_id = request.form.get('categoria_id')

    if not producto:
        flash("Debes especificar el producto pulido.", "warning")
        return redirect(url_for("dashboard"))

    try:
        cantidad = int(cantidad_str)
        if cantidad <= 0:
            raise ValueError
    except ValueError:
        flash("Cantidad inválida.", "warning")
        return redirect(url_for("dashboard"))

    reg = RegistroPulido(
        fecha=date.today(),
        usuario_id=current_user.id,
        producto=producto,
        acabado=acabado,
        cantidad=cantidad,
        observaciones=observaciones,
        categoria_id=int(categoria_id),
        estado="pulido"
    )

    db.session.add(reg)
    db.session.commit()
    flash("Pulido registrado ✅", "success")
    return redirect(url_for("dashboard"))



@app.route('/inventario', endpoint='inventario')
@login_required
def inventario():
    categoria_id = request.args.get('categoria', type=int)
    categorias = CategoriaInsumo.query.order_by(CategoriaInsumo.nombre.asc()).all()

    query = Insumo.query
    if categoria_id:
        query = query.filter_by(categoria_id=categoria_id)

    insumos = query.order_by(Insumo.nombre.asc()).all()
    nombres = [i.nombre for i in insumos]
    cantidades = [i.cantidad_actual for i in insumos]

    return render_template(
        'inventario.html',
        insumos=insumos,
        nombres=nombres,
        cantidades=cantidades,
        categorias=categorias
    )


@app.route('/insumos', methods=['GET', 'POST'])
@login_required
@require_roles('admin')
def insumos_create():
    categorias = CategoriaInsumo.query.order_by(CategoriaInsumo.nombre.asc()).all()

    if request.method == 'POST':
        nombre = request.form.get('nombre')
        unidad = request.form.get('unidad')
        cantidad = request.form.get('cantidad')
        bodega = request.form.get('bodega')
        categoria_nombre = request.form.get('categoria_nombre')
        categoria_id = request.form.get('categoria_id')

        # Si el usuario escribió una nueva categoría
        if categoria_nombre:
            nueva_cat = CategoriaInsumo(nombre=categoria_nombre.strip())
            db.session.add(nueva_cat)
            db.session.commit()
            categoria_id = nueva_cat.id

        nuevo = Insumo(
            nombre=nombre.strip(),
            unidad=unidad.strip(),
            cantidad_actual=float(cantidad),
            bodega=bodega.strip(),
            categoria_id=int(categoria_id) if categoria_id else None
        )
        db.session.add(nuevo)
        db.session.commit()
        flash('Insumo creado ✅', 'success')
        return redirect(url_for('inventario'))

    return render_template('insumos.html', categorias=categorias)

@app.route('/insumos/<int:insumo_id>/editar', methods=['GET', 'POST'])
@login_required
@require_roles('admin')
def editar_insumo(insumo_id):
    insumo = Insumo.query.get_or_404(insumo_id)
    categorias = CategoriaInsumo.query.order_by(CategoriaInsumo.nombre.asc()).all()

    if request.method == 'POST':
        insumo.nombre = request.form.get('nombre', insumo.nombre).strip()
        insumo.unidad = request.form.get('unidad', insumo.unidad).strip()
        insumo.cantidad_actual = float(request.form.get('cantidad', insumo.cantidad_actual))
        insumo.bodega = request.form.get('bodega', insumo.bodega).strip()

        # Actualizar categoría (existente o nueva)
        categoria_nombre = request.form.get('categoria_nombre')
        categoria_id = request.form.get('categoria_id')

        if categoria_nombre:
            nueva_cat = CategoriaInsumo(nombre=categoria_nombre.strip())
            db.session.add(nueva_cat)
            db.session.commit()
            insumo.categoria_id = nueva_cat.id
        elif categoria_id:
            insumo.categoria_id = int(categoria_id)

        db.session.commit()
        flash('Insumo actualizado ✅', 'success')
        return redirect(url_for('inventario'))

    return render_template('editar_insumo.html', insumo=insumo, categorias=categorias)


@app.route('/insumos/<int:insumo_id>/eliminar', methods=['POST'])
@login_required
@require_roles('admin')
def eliminar_insumo(insumo_id):
    ins = Insumo.query.get_or_404(insumo_id)
    Movimiento.query.filter_by(insumo_id=ins.id).delete()
    db.session.delete(ins)
    db.session.commit()
    flash('Insumo eliminado 🗑️', 'warning')
    return redirect(url_for('inventario'))

@app.route('/movimiento', methods=['GET', 'POST'])
@login_required
def movimiento_insumo():

    insumos = Insumo.query.order_by(Insumo.nombre.asc()).all()
    if request.method == 'POST':
        tipo = request.form.get('tipo')
        cantidad = float(request.form.get('cantidad'))
        insumo_id = int(request.form.get('insumo'))
        insumo = Insumo.query.get(insumo_id)

        if not insumo:
            flash('Insumo no encontrado', 'danger')
            return render_template('movimiento.html', insumos=insumos)

        if tipo == 'Entrada':
            insumo.cantidad_actual += cantidad
        elif tipo == 'Salida':
            if insumo.cantidad_actual >= cantidad:
                insumo.cantidad_actual -= cantidad
            else:
                flash('No hay suficiente stock', 'danger')
                return render_template('movimiento.html', insumos=insumos)
        else:
            flash('Tipo inválido', 'danger')
            return render_template('movimiento.html', insumos=insumos)

        mov = Movimiento(tipo=tipo, cantidad=cantidad, insumo=insumo)
        db.session.add(mov)
        db.session.commit()
        flash('Movimiento registrado ✅', 'success')
        return redirect(url_for('inventario'))
    return render_template('movimiento.html', insumos=insumos)

@app.route('/historial')
@login_required
def historial_insumos():
    insumo_id = request.args.get('insumo_id', type=int)
    tipo = request.args.get('tipo')
    desde = request.args.get('desde')
    hasta = request.args.get('hasta')

    q = db.session.query(Movimiento, Insumo).join(Insumo, Movimiento.insumo_id == Insumo.id)

    if insumo_id:
        q = q.filter(Movimiento.insumo_id == insumo_id)
    if tipo in ('Entrada', 'Salida'):
        q = q.filter(Movimiento.tipo == tipo)
    if desde:
        try:
            d = datetime.strptime(desde, "%Y-%m-%d")
            q = q.filter(Movimiento.fecha >= d)
        except ValueError:
            flash('Fecha "desde" inválida', 'danger')
    if hasta:
        try:
            h = datetime.strptime(hasta, "%Y-%m-%d")
            h = h.replace(hour=23, minute=59, second=59)
            q = q.filter(Movimiento.fecha <= h)
        except ValueError:
            flash('Fecha "hasta" inválida', 'danger')

    movs = q.order_by(Movimiento.fecha.desc()).all()
    insumos = Insumo.query.order_by(Insumo.nombre.asc()).all()

    return render_template('historial.html', movs=movs, insumos=insumos,
                           sel_insumo_id=insumo_id or '',
                           sel_tipo=tipo or '',
                           sel_desde=desde or '',
                           sel_hasta=hasta or '')

@app.route('/usuarios')
@login_required
@require_roles('admin')
def usuarios():
    usuarios = Usuario.query.order_by(Usuario.nombre.asc()).all()
    return render_template('usuarios.html', usuarios=usuarios)

@app.route('/usuarios/<int:user_id>/rol', methods=['POST'])
@login_required
@require_roles('admin')
def usuarios_cambiar_rol(user_id):
    nuevo_rol = request.form.get('rol')
    if nuevo_rol not in ['admin', 'operario']:
        flash('Rol inválido', 'danger')
        return redirect(url_for('usuarios'))
    u = Usuario.query.get_or_404(user_id)
    u.rol = nuevo_rol
    db.session.commit()
    flash('Rol actualizado ✅', 'success')
    return redirect(url_for('usuarios'))

@app.route('/export/inventario.xlsx')
@login_required
@require_roles('admin')
def export_inventario():
    insumos = Insumo.query.order_by(Insumo.nombre.asc()).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"
    headers = ["Nombre", "Unidad", "Cantidad actual", "Bodega"]
    ws.append(headers)
    for i in insumos:
        ws.append([i.nombre, i.unidad, i.cantidad_actual, i.bodega])
    for col in range(1, len(headers)+1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True,
                     download_name='inventario.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/export/movimientos.xlsx')
@login_required
@require_roles('admin')
def export_movs_insumos():
    movs = (db.session.query(Movimiento, Insumo)
            .join(Insumo, Movimiento.insumo_id == Insumo.id)
            .order_by(Movimiento.fecha.desc())
            .all())
    wb = Workbook()
    ws = wb.active
    ws.title = "Movimientos"
    headers = ["Fecha", "Insumo", "Tipo", "Cantidad", "Unidad", "Bodega"]
    ws.append(headers)
    for m, i in movs:
        ws.append([fmt_local(m.fecha, "%Y-%m-%d %H:%M"), i.nombre, m.tipo, m.cantidad, i.unidad, i.bodega])
    for col in range(1, len(headers)+1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True,
                     download_name='movimientos.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/productos_por_categoria/<int:categoria_id>/tabla')
@login_required
def productos_por_categoria_tabla(categoria_id):
    if categoria_id == 0:  # caso especial "sin categoría"
        productos = Producto.query.filter(Producto.categoria_id.is_(None)).order_by(Producto.nombre.asc()).all()
    else:
        productos = Producto.query.filter_by(categoria_id=categoria_id).order_by(Producto.nombre.asc()).all()
    
    data = [
        {
            'id': p.id,
            'nombre': p.nombre,
            'acabado': p.acabado,
            'cantidad': p.cantidad_actual,
            'bodega': p.bodega,
            'categoria': p.categoria_producto.nombre if p.categoria_producto else '-'
        }
        for p in productos
    ]

    return jsonify(data)



@app.route('/productos_por_categoria/<int:categoria_id>', methods=['GET', 'POST'])
@login_required
def productos_por_categoria_vista(categoria_id):
    categorias = CategoriaProduccion.query.order_by(CategoriaProduccion.nombre.asc()).all()
    BODEGA_ACTUAL = "Tocancipa"   # ← Aquí defines tu bodega

    if categoria_id == 0:
        categoria_nombre = "Sin categoría"
        productos = (
            Producto.query
            .filter(Producto.categoria_id.is_(None))
            .filter(Producto.bodega == BODEGA_ACTUAL)
            .order_by(Producto.nombre.asc())
            .all()
        )
    else:
        categoria = CategoriaProduccion.query.get_or_404(categoria_id)
        categoria_nombre = categoria.nombre
        productos = (
            Producto.query
            .filter(Producto.categoria_id == categoria_id)
            .filter(Producto.bodega == BODEGA_ACTUAL)
            .order_by(Producto.nombre.asc())
            .all()
        )

    # Si cambian categoría desde la vista
    if request.method == 'POST':
        producto_id = request.form.get('producto_id')
        nueva_categoria = request.form.get('nueva_categoria')
        if producto_id and nueva_categoria:
            producto = Producto.query.get(producto_id)
            producto.categoria_id = int(nueva_categoria) if nueva_categoria != "0" else None
            db.session.commit()
            flash('Categoría actualizada correctamente', 'success')
            return redirect(request.url)

    return render_template(
        'productos_por_categoria.html',
        categoria_nombre=categoria_nombre,
        productos=productos,
        categorias=categorias,
        categoria_id=categoria_id
    )


@app.route('/produccion', endpoint='produccion')
@login_required
def produccion():
    q = (request.args.get('q') or '').strip()
    categoria_id = request.args.get('categoria', type=int)

    categorias = CategoriaProduccion.query.order_by(CategoriaProduccion.nombre.asc()).all()

    query = Producto.query

    # 🔍 Buscador
    if q:
        like = f"%{q}%"
        query = query.filter(
            or_(
                Producto.nombre.ilike(like),
                Producto.acabado.ilike(like),
                Producto.bodega.ilike(like),
            )
        )

    # 🔥 FILTRO POR BODEGA DEL LÍDER
    query = query.filter(Producto.bodega == "Tocancipa")

    # 🔥 Filtro por categoría
    if categoria_id:
        query = query.filter(Producto.categoria_id == categoria_id)

    productos = query.order_by(Producto.nombre.asc()).all()

    # --- totales por categoría (solo de Tocancipá)
    totales_categoria = (
        db.session.query(
            CategoriaProduccion.id,
            CategoriaProduccion.nombre,
            func.coalesce(func.sum(Producto.cantidad_actual), 0).label('total')
        )
        .select_from(CategoriaProduccion)
        .outerjoin(Producto, CategoriaProduccion.id == Producto.categoria_id)
        .filter(Producto.bodega == "Tocancipa")  # 👈 importante
        .group_by(CategoriaProduccion.id)
        .order_by(CategoriaProduccion.nombre.asc())
        .all()
    )

    # sin categoría (solo Tocancipá)
    sin_categoria_total = (
        db.session.query(func.coalesce(func.sum(Producto.cantidad_actual), 0))
        .filter(Producto.categoria_id.is_(None))
        .filter(Producto.bodega == "Tocancipa")
        .scalar()
    )

    if sin_categoria_total > 0:
        totales_categoria.append((0, "Sin categoría", sin_categoria_total))

    db.session.expire_all()

    return render_template(
        'produccion.html',
        productos=productos,
        categorias=categorias,
        totales_categoria=totales_categoria,
        q=q
    )

@app.route('/produccion/<int:id_categoria>')
@login_required
def produccion_categoria(id_categoria):
    categoria = CategoriaProduccion.query.get_or_404(id_categoria)
    search = request.args.get('search', '')

    productos = (
        Producto.query
        .filter(Producto.categoria_id == id_categoria)
        .filter(Producto.bodega == "Tocancipa")     # 👈 FILTRO IMPORTANTE
        .filter(
            (Producto.nombre.ilike(f'%{search}%')) |
            (Producto.acabado.ilike(f'%{search}%'))
        )
        .order_by(Producto.nombre.asc())
        .all()
    )

    return render_template(
        'productos_por_categoria.html', 
        categoria=categoria, 
        productos=productos, 
        search=search
    )




@app.route('/productos', methods=['GET', 'POST'], endpoint='crear_producto')
@login_required
@require_roles('admin')
def crear_producto():
    categorias = CategoriaProduccion.query.order_by(CategoriaProduccion.nombre.asc()).all()

    if request.method == 'POST':
        nombre = (request.form.get('nombre') or '').strip()
        acabado = (request.form.get('acabado') or '').strip()
        cantidad_str = (request.form.get('cantidad') or '').strip()
        bodega = (request.form.get('bodega') or '').strip()
        categoria_nombre = (request.form.get('categoria_nombre') or '').strip()
        categoria_id = request.form.get('categoria_id')

        # Validaciones
        if not nombre or not acabado or not cantidad_str or not bodega:
            flash('Faltan datos.', 'danger')
            return render_template('producto_nuevo.html',
                                   nombre=nombre, acabado=acabado,
                                   cantidad=cantidad_str, bodega=bodega,
                                   categorias=categorias)

        cantidad_str = cantidad_str.replace(',', '.')
        try:
            cantidad = float(cantidad_str)
        except ValueError:
            flash('Cantidad inválida.', 'danger')
            return render_template('producto_nuevo.html',
                                   nombre=nombre, acabado=acabado,
                                   cantidad=cantidad_str, bodega=bodega,
                                   categorias=categorias)

        # Si el usuario escribió una nueva categoría
        if categoria_nombre:
            nueva_cat = CategoriaProduccion(nombre=categoria_nombre)
            db.session.add(nueva_cat)
            db.session.commit()
            categoria_id = nueva_cat.id

        nuevo = Producto(
            nombre=nombre,
            acabado=acabado,
            cantidad_actual=cantidad,
            bodega=bodega,
            categoria_id=int(categoria_id) if categoria_id else None
        )
        db.session.add(nuevo)
        db.session.commit()
        flash('Producto creado ✅', 'success')
        return go_produccion()

    return render_template('producto_nuevo.html', categorias=categorias)


# --- EDITAR PRODUCTO ---
@app.route('/producto/<int:producto_id>/editar', methods=['GET', 'POST'])
@login_required
@require_roles('admin')
def producto_editar(producto_id):
    producto = Producto.query.get_or_404(producto_id)
    categorias = CategoriaProduccion.query.order_by(CategoriaProduccion.nombre.asc()).all()

    if request.method == 'POST':
        producto.nombre = request.form['nombre']
        producto.acabado = request.form['acabado']
        producto.cantidad_actual = float(request.form['cantidad_actual'])
        producto.bodega = request.form['bodega']

        categoria_id = request.form.get('categoria_id')
        if categoria_id:
            producto.categoria_id = int(categoria_id)
        else:
            producto.categoria_id = None

        db.session.commit()
        flash('Producto actualizado correctamente ✅', 'success')
        return redirect(url_for('produccion'))

    return render_template('editar_producto.html', producto=producto, categorias=categorias)


@app.route('/productos/<int:producto_id>/eliminar', methods=['POST'])
@login_required
@require_roles('admin')
def producto_eliminar(producto_id):
    p = Producto.query.get_or_404(producto_id)
    ProdMovimiento.query.filter_by(producto_id=p.id).delete()
    db.session.delete(p)
    db.session.commit()
    flash('Producto eliminado 🗑️', 'warning')
    return go_produccion()


@app.route('/movimiento-produccion', methods=['GET', 'POST'])
@login_required
def movimiento_produccion():
    # Helpers para llenar los combos
    def _dropdowns(nombre_pref=''):
        productos_all = Producto.query.order_by(Producto.nombre.asc()).all()
        nombres = sorted({p.nombre for p in productos_all})
        nombre_sel = nombre_pref or (nombres[0] if nombres else '')
        acabados = sorted({p.acabado for p in productos_all if p.nombre == nombre_sel})
        # bodegas para transferencias
        bodegas_db = [b[0] for b in db.session.query(Producto.bodega).distinct().all() if b[0]]
        bodegas = sorted({*BODEGAS_FIJAS, *bodegas_db})
        return productos_all, nombres, nombre_sel, acabados, bodegas

    if request.method == 'GET':
        nombre_sel = (request.args.get('nombre') or '').strip()
        productos_all, nombres, nombre_sel, acabados, bodegas = _dropdowns(nombre_sel)
        return render_template(
            'movimiento_produccion.html',
            nombres=nombres, nombre_sel=nombre_sel, acabados=acabados,
            bodegas=bodegas
        )

    # POST
    # Traemos nombre + acabado en vez de product_id
    nombre = (request.form.get('nombre') or '').strip()
    acabado = (request.form.get('acabado') or '').strip()
    tipo = (request.form.get('tipo') or '').strip()
    cantidad = float(request.form.get('cantidad') or 0)

    p = Producto.query.filter_by(nombre=nombre, acabado=acabado).first()
    if not p:
        productos_all, nombres, nombre_sel, acabados, bodegas = _dropdowns(nombre)
        flash('Producto/acabado no encontrado.', 'danger')
        return render_template(
            'movimiento_produccion.html',
            nombres=nombres, nombre_sel=nombre_sel, acabados=acabados,
            bodegas=bodegas
        )

    productos_all, nombres, nombre_sel, acabados, bodegas = _dropdowns(nombre)

    # --- Transferencia (bodega->bodega) ---
    if tipo == 'Transferencia':
        destino = (request.form.get('bodega_destino') or '').strip()

        if not destino or destino == p.bodega:
            flash('Selecciona una bodega destino distinta a la de origen.', 'warning')
            return render_template('movimiento_produccion.html',
                                   nombres=nombres, nombre_sel=nombre_sel, acabados=acabados, bodegas=bodegas)

        if cantidad <= 0:
            flash('Cantidad inválida.', 'danger')
            return render_template('movimiento_produccion.html',
                                   nombres=nombres, nombre_sel=nombre_sel, acabados=acabados, bodegas=bodegas)

        if p.cantidad_actual < cantidad:
            flash('No hay suficiente stock del producto', 'danger')
            return render_template('movimiento_produccion.html',
                                   nombres=nombres, nombre_sel=nombre_sel, acabados=acabados, bodegas=bodegas)
        

        # 1) Origen: salida
        p.cantidad_actual -= cantidad
        mov_origen = ProdMovimiento(
            tipo='Salida',
            cantidad=cantidad,
            producto=p,
            fecha=datetime.utcnow()
        )
        
        db.session.add(mov_origen)
        db.session.flush()

        db.session.execute(
            text("UPDATE prod_movimiento SET bodega=:b WHERE id=:id"),
            {"b": p.bodega, "id": mov_origen.id}
        )

        # 2) Destino
        dest = (Producto.query
                .filter_by(nombre=p.nombre, acabado=p.acabado, bodega=destino)
                .first())
        if not dest:
            dest = Producto(
                nombre=p.nombre, acabado=p.acabado, cantidad_actual=0, bodega=destino, categoria_id=p.categoria_id
            )
            db.session.add(dest)
            db.session.flush()

        dest.cantidad_actual += cantidad

        mov_dest = ProdMovimiento(
            tipo='Entrada',
            cantidad=cantidad,
            producto=dest,
            fecha=datetime.utcnow()
        )
        db.session.add(mov_dest)
        db.session.flush()

        db.session.execute(
            text("UPDATE prod_movimiento SET bodega=:b WHERE id=:id"),
            {"b": destino, "id": mov_dest.id}
        )
        
        # 3) Registrar la transferencia completa en una sola fila
        t = Transferencia(
        producto_nombre=p.nombre,
        acabado=p.acabado,
        cantidad=cantidad,
        origen=p.bodega,
        destino=destino,
        fecha=datetime.utcnow()
        )
        db.session.add(t)


        db.session.commit()
        flash(f"Transferencia realizada: {cantidad} de “{p.nombre}” ({p.acabado}) de {p.bodega} → {destino}", 'success')
        return redirect(url_for('movimiento_produccion'))


    # --- Entrada / Salida simples ---
    if cantidad <= 0:
        flash('Cantidad inválida.', 'danger')
        return render_template('movimiento_produccion.html',
                               nombres=nombres, nombre_sel=nombre_sel, acabados=acabados, bodegas=bodegas)

    if tipo == 'Entrada':
        p.cantidad_actual += cantidad
    elif tipo == 'Salida':
        if p.cantidad_actual < cantidad:
            flash('No hay suficiente stock del producto', 'danger')
            return render_template('movimiento_produccion.html',
                                   nombres=nombres, nombre_sel=nombre_sel, acabados=acabados, bodegas=bodegas)
        p.cantidad_actual -= cantidad
    else:
        flash('Tipo inválido', 'danger')
        return render_template('movimiento_produccion.html',
                               nombres=nombres, nombre_sel=nombre_sel, acabados=acabados, bodegas=bodegas)

    mov = ProdMovimiento(
        tipo=tipo, cantidad=cantidad, producto=p, fecha=datetime.utcnow()
    )
    db.session.add(mov)
    db.session.flush()

    db.session.execute(
        text("UPDATE prod_movimiento SET bodega=:b WHERE id=:id"),
        {"b": p.bodega, "id": mov.id}
    )

    db.session.commit()
    flash('Movimiento registrado ✅', 'success')
    return redirect(url_for('movimiento_produccion'))


@app.route('/historial-produccion')
@login_required
def historial_produccion():
    producto_id = request.args.get('producto_id', type=int)
    tipo = request.args.get('tipo')
    desde = request.args.get('desde')
    hasta = request.args.get('hasta')

    q = db.session.query(ProdMovimiento, Producto).join(Producto, ProdMovimiento.producto_id == Producto.id)

    if producto_id:
        q = q.filter(ProdMovimiento.producto_id == producto_id)
    if tipo in ('Entrada', 'Salida'):
        q = q.filter(ProdMovimiento.tipo == tipo)
    if desde:
        try:
            d = datetime.strptime(desde, "%Y-%m-%d")
            q = q.filter(ProdMovimiento.fecha >= d)
        except ValueError:
            flash('Fecha "desde" inválida', 'danger')
    if hasta:
        try:
            h = datetime.strptime(hasta, "%Y-%m-%d")
            h = h.replace(hour=23, minute=59, second=59)
            q = q.filter(ProdMovimiento.fecha <= h)
        except ValueError:
            flash('Fecha "hasta" inválida', 'danger')

    movs = q.order_by(ProdMovimiento.fecha.desc()).all()
    productos = Producto.query.order_by(Producto.nombre.asc()).all()

    return render_template('historial_produccion.html', movs=movs, productos=productos,
                           sel_producto_id=producto_id or '',
                           sel_tipo=tipo or '',
                           sel_desde=desde or '',
                           sel_hasta=hasta or '')
app.add_url_rule('/produccion',
    endpoint='produccion_view',
    view_func=app.view_functions['produccion']
)

app.add_url_rule( '/movimiento-produccion',
    endpoint='movimiento_produccion_view',
    view_func=app.view_functions['movimiento_produccion']
)

app.add_url_rule(
    '/historial-produccion',
    endpoint='historial_produccion_view',
    view_func=app.view_functions['historial_produccion']
)
from flask import redirect, url_for

def go_produccion():
    try:
        return redirect(url_for('produccion'))       
    except Exception:
        try:
            return redirect(url_for('produccion_view')) 
        except Exception:
            return redirect('/produccion')       

@app.route('/transferencias')
@login_required
def transferencias():
    destino = request.args.get('destino', '').strip()
    fecha_desde = request.args.get('desde', '')
    fecha_hasta = request.args.get('hasta', '')

    q = Transferencia.query.order_by(Transferencia.fecha.desc())

    # FILTRAR POR DESTINO
    if destino:
        q = q.filter(Transferencia.destino == destino)

    # FILTRAR POR FECHA DESDE
    if fecha_desde:
        try:
            d = datetime.strptime(fecha_desde, "%Y-%m-%d")
            q = q.filter(Transferencia.fecha >= d)
        except:
            pass

    # FILTRAR POR FECHA HASTA
    if fecha_hasta:
        try:
            h = datetime.strptime(fecha_hasta, "%Y-%m-%d")
            h = h.replace(hour=23, minute=59, second=59)
            q = q.filter(Transferencia.fecha <= h)
        except:
            pass

    todas = q.all()

    # Para evitar destinos repetidos
    destinos_unicos = sorted({t.destino for t in Transferencia.query.all()})

    return render_template(
        'transferencias.html',
        transferencias=todas,
        sel_destino=destino,
        sel_desde=fecha_desde,
        sel_hasta=fecha_hasta,
        destinos=destinos_unicos
    )

@app.route('/exportar_transferencias')
@login_required
def exportar_transferencias():
    destino = request.args.get('destino', '').strip()
    fecha = request.args.get('fecha', '').strip()

    q = Transferencia.query.order_by(Transferencia.fecha.desc())

    if destino:
        q = q.filter(Transferencia.destino == destino)

    if fecha:
        try:
            d = datetime.strptime(fecha, "%Y-%m-%d")
            ini = datetime.combine(d.date(), datetime.min.time())
            fin = datetime.combine(d.date(), datetime.max.time())
            q = q.filter(Transferencia.fecha >= ini, Transferencia.fecha <= fin)
        except:
            pass

    datos = q.all()

    # Crear Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Transferencias"

    headers = ["Fecha", "Producto", "Acabado", "Cantidad", "Origen", "Destino"]
    ws.append(headers)

    for t in datos:
        ws.append([
            t.fecha.strftime("%Y-%m-%d %H:%M"),
            t.producto_nombre,
            t.acabado,
            t.cantidad,
            t.origen,
            t.destino
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name='transferencias_filtradas.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

              

@app.route('/movimientos', endpoint='movimientos')
@login_required
def movimientos():
    movs = (
        db.session.query(Movimiento, Insumo)
        .join(Insumo, Movimiento.insumo_id == Insumo.id)
        .order_by(Movimiento.fecha.desc())
        .limit(200)
        .all()
    )
    return render_template('movimientos_recientes.html', movs=movs)



@app.route('/registrar', methods=['GET', 'POST'])
def registrar():
    hay_usuarios = Usuario.query.count() > 0

    if hay_usuarios:
        if not current_user.is_authenticated:
            flash('Inicia sesión para crear usuarios.', 'warning')
            return redirect(url_for('login'))
        if current_user.rol != 'admin':
            flash('⛔ Solo un administrador puede crear usuarios.', 'danger')
            return redirect(url_for('dashboard'))

    if request.method == 'POST':
        nombre = (request.form.get('nombre') or '').strip()
        email  = (request.form.get('email') or '').strip()
        password = (request.form.get('password') or '')

        if not nombre or not email or not password:
            flash('Completa todos los campos.', 'danger')
            return render_template('registrar.html')

        if Usuario.query.filter_by(email=email).first():
            flash('Ese correo ya existe.', 'danger')
            return render_template('registrar.html')

        if not hay_usuarios:
            rol = 'admin'
        else:
            rol_form = (request.form.get('rol') or 'operario').strip().lower()
            rol = 'admin' if rol_form == 'admin' else 'operario'

        nuevo = Usuario(
            nombre=nombre,
            email=email,
            password=generate_password_hash(password),
            rol=rol
        )
        db.session.add(nuevo)
        db.session.commit()

        if not hay_usuarios:
            flash('Primer usuario creado como ADMIN. Inicia sesión.', 'success')
            return redirect(url_for('login'))

        flash('Usuario creado ✅', 'success')
        return redirect(url_for('usuarios'))

    return render_template('registrar.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form.get('email') or ''
        password = request.form.get('password') or ''
        u = Usuario.query.filter_by(email=email).first()
        if u and check_password_hash(u.password, password):
            login_user(u)
            flash(f'Bienvenido, {u.nombre} 👋', 'success')
            return redirect(request.args.get('next') or url_for('dashboard'))
        flash('Credenciales incorrectas ❌', 'danger')
        return render_template('login.html'), 401
    return render_template('login.html')

@app.route("/_routes")
def _routes():
    out = []
    for r in app.url_map.iter_rules():
        out.append(f"{r.endpoint:30s} -> {r.rule}")
    return "<pre>" + "\n".join(sorted(out)) + "</pre>"

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/categorias', methods=['GET', 'POST'])
@login_required
@require_roles('admin')
def categorias():
    if request.method == 'POST':
        nombre = (request.form.get('nombre') or '').strip()
        descripcion = (request.form.get('descripcion') or '').strip()
        if not nombre:
            flash('El nombre de la categoría es obligatorio.', 'warning')
        else:
            nueva = CategoriaInsumo(nombre=nombre, descripcion=descripcion)
            db.session.add(nueva)
            db.session.commit()
            flash('Categoría creada ✅', 'success')
        return redirect(url_for('categorias'))

    categorias = CategoriaInsumo.query.order_by(CategoriaInsumo.nombre.asc()).all()
    return render_template('categorias.html', categorias=categorias)
from flask import send_file
import os

@app.route('/descargar_bd')
@login_required
@require_roles('admin')
def descargar_bd():
    db_path = "/var/data/database.db"  # Ruta real del archivo activo en Render
    if os.path.exists(db_path):
        return send_file(db_path, as_attachment=True, download_name='database.db')
    else:
        return "No se encontró la base de datos en el servidor.", 404
    
    

if __name__ == '__main__':
    app.run(debug=True)
